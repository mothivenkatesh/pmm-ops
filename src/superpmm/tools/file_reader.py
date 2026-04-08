"""File reader for uploaded documents (PRDs, PRFAQs, call notes, etc.)."""

import os
import json


def read_uploaded_file(file_path: str) -> str:
    """Read and extract text from uploaded files.

    Supports: .txt, .md, .pdf, .docx, .xlsx
    """
    if not os.path.exists(file_path):
        return f"File not found: {file_path}"

    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".txt", ".md"):
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()

    elif ext == ".pdf":
        return _read_pdf(file_path)

    elif ext == ".docx":
        return _read_docx(file_path)

    elif ext in (".xlsx", ".csv"):
        return _read_spreadsheet(file_path)

    else:
        return f"Unsupported file type: {ext}. Supported: .txt, .md, .pdf, .docx, .xlsx, .csv"


def _read_pdf(path: str) -> str:
    try:
        import fitz  # pymupdf
        doc = fitz.open(path)
        text = []
        for page in doc:
            text.append(page.get_text())
        return "\n\n".join(text)
    except ImportError:
        return "PDF reading requires pymupdf: pip install pymupdf"


def _read_docx(path: str) -> str:
    try:
        from docx import Document
        doc = Document(path)
        text = []
        for para in doc.paragraphs:
            if para.text.strip():
                text.append(para.text)
        return "\n\n".join(text)
    except ImportError:
        return "DOCX reading requires python-docx: pip install python-docx"


def _read_spreadsheet(path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text.append(f"## Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    text.append(" | ".join(cells))
        return "\n".join(text)
    except ImportError:
        return "XLSX reading requires openpyxl: pip install openpyxl"
